# =============================================================================
# Module imports
# =============================================================================

import os
from collections import Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font

# =============================================================================
# Function imports
# =============================================================================

from ParseRackPrint import generate_rack_summary_file, generate_rack_folder

# =============================================================================
# Constants
# =============================================================================

customer_group_file=r"J:\Duplicate Rack Summaries\customer_groups.xlsx"
rack_summaries_folder = r"P:\Public\Past 7 days RackPrints"

# =============================================================================
# Helper functions
# =============================================================================

def highlight_cell(cell, color="FFFF00"):
    """
    Highlight a cell with the specified color (default is yellow).

    Args:
        cell: The cell to highlight.
        color (str): The color to use for highlighting.

    Returns:
        None
    """
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill

def get_delivered_glass(file_path, runs):

    runs = [run.split('-')[0] for run in runs]

    # Load workbook and first sheet
    wb = load_workbook(file_path, data_only=False, read_only=False)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]

    col_index = {header: i for i, header in enumerate(headers)}

    new_order = ['Item Reference', 'Order No', 'Delivery Address', 'Product', 'Specification', 'Rack Number']

    new_data = []
    for row in ws.iter_rows(values_only=True):
        run = row[col_index['Despatch Run']]
        if str(run) in runs:
            new_data.append([row[col_index[h]] for h in new_order])

    wb.remove(ws)
    ws = wb.create_sheet("Reordered")  

    # Write rearranged data
    for row in new_data:
        ws.append(row)

    delivered = {}

    for row in ws.iter_rows(min_row=3):

        piece_id = row[0].value  # Column A (index 0)

        row_data = []
        if piece_id is not None:
            for cell in row:
                row_data.append(cell.value)
            delivered[piece_id] = row_data

    return delivered

def get_manifest_glass(file_path):

    # Load workbook and first sheet
    wb = load_workbook(file_path, data_only=False, read_only=False)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]

    col_index = {header: i for i, header in enumerate(headers)}

    new_order = ['Order_No', 'Customer', 'Product', 'Height', 'Width']

    new_data = []
    for row in ws.iter_rows(values_only=True):
        new_data.append([row[col_index[h]] for h in new_order])

    wb.remove(ws)
    ws = wb.create_sheet("Reordered")  

    # Write rearranged data
    for row in new_data:
        ws.append(row)

    ws.cell(row=1, column=len(new_order), value="Size")

    manifested = []

    for row in ws.iter_rows(min_row=3):

        row_data = []
        for cell in row:
            row_data.append(cell.value)

        row_data.append(f"{row_data[3]}x{row_data[4]}")  # Add Size as "Height x Width"

        row_data.pop(3)
        row_data.pop(3)

        manifested.append(row_data)

    manifested.pop(-1)

    return manifested

def get_rack_summary_glass(folder_path):

    rack_summary_path, header = generate_rack_summary_file(folder_path)

    # Load workbook and first sheet
    wb = load_workbook(rack_summary_path, data_only=False, read_only=False)
    ws = wb.active

    rack_summaries = {}

    for row in ws.iter_rows(min_row=3):

        piece_id = row[0].value  # Column A (index 0)

        row_data = []
        if piece_id is not None:
            for cell in row:
                row_data.append(cell.value)
            rack_summaries[piece_id] = row_data

    return rack_summaries, header

# =============================================================================
# Main functions
# =============================================================================

def compare_rack_scanned_reports(rack_folder_path, scanned_file_path, runs):

    rack_summary_glass, header = get_rack_summary_glass(rack_folder_path)
    delivered = get_delivered_glass(scanned_file_path, runs)

    matching_keys = set(rack_summary_glass.keys()) & set(delivered.keys())
    matching_values = [rack_summary_glass.get(key) for key in matching_keys]
    matching_piece_ids = [matching_value[0] for matching_value in matching_values if matching_value]

    wb = Workbook()

    ws_rack = wb.create_sheet("Rack Summaries", 0)
    ws_scanned = wb.create_sheet("Scanned Glass", 1)
    ws_rack_discrepancies = wb.create_sheet("Rack Discrepancies", 2)
    ws_scanned_discrepancies = wb.create_sheet("Scanned Discrepancies", 3)

    ws_rack_discrepancies['A1'] = "Rack vs Scanned: Rack Discrepancies"
    ws_rack_discrepancies['A1'].font = Font(bold=True, size=14)

    for value in header:
        col_index = header.index(value) + 1
        ws_rack_discrepancies.cell(row=2, column=col_index, value=value)
        ws_rack_discrepancies.cell(row=2, column=col_index).font = Font(bold=True)
        highlight_cell(ws_rack_discrepancies.cell(row=2, column=col_index), color="FFADD8E6")
    ws_rack_discrepancies.cell(row=2, column=len(header) + 1, value="Manifest Status")
    ws_rack_discrepancies.cell(row=2, column=len(header) + 1).font = Font(bold=True)
    highlight_cell(ws_rack_discrepancies.cell(row=2, column=len(header) + 1), color="FFADD8E6")

    ws_scanned_discrepancies['A1'] = "Rack vs Scanned: Scanned Discrepancies"
    ws_scanned_discrepancies['A1'].font = Font(bold=True, size=14)

    for value in header:
        col_index = header.index(value) + 1
        ws_scanned_discrepancies.cell(row=2, column=col_index, value=value) 
        ws_scanned_discrepancies.cell(row=2, column=col_index).font = Font(bold=True)
        highlight_cell(ws_scanned_discrepancies.cell(row=2, column=col_index), color="FFADD8E6")
    ws_scanned_discrepancies.cell(row=2, column=len(header) + 1, value="Manifest Status")
    ws_scanned_discrepancies.cell(row=2, column=len(header) + 1).font = Font(bold=True)
    highlight_cell(ws_scanned_discrepancies.cell(row=2, column=len(header) + 1), color="FFADD8E6")

    ws_rack['A1'] = "Rack Summary Glass"
    ws_rack['A1'].font = Font(bold=True, size=14)
    for value in header:
        col_index = header.index(value) + 1
        ws_rack.cell(row=2, column=col_index, value=value)   
        ws_rack.cell(row=2, column=col_index).font = Font(bold=True)
        highlight_cell(ws_rack.cell(row=2, column=col_index), color="FFADD8E6")
    ws_rack.cell(row=2, column=len(header) + 1, value="Match Status (Compared to Scanned)")
    ws_rack.cell(row=2, column=len(header) + 1).font = Font(bold=True)
    highlight_cell(ws_rack.cell(row=2, column=len(header) + 1), color="FFADD8E6")
    for index, value in enumerate(rack_summary_glass.values(), start=3):
        if matching_piece_ids and value[0] in matching_piece_ids:
            for col_index, cell_value in enumerate(value, start=1):
                ws_rack.cell(row=index, column=col_index, value=cell_value)
                highlight_cell(ws_rack.cell(row=index, column=col_index), color="FF98FB98")
            ws_rack.cell(row=index, column=len(header) + 1, value="Match")
            highlight_cell(ws_rack.cell(row=index, column=len(header) + 1), color="FF98FB98")
        else:
            for col_index, cell_value in enumerate(value, start=1):
                ws_rack.cell(row=index, column=col_index, value=cell_value)
                highlight_cell(ws_rack.cell(row=index, column=col_index), color="FFE78587")
            ws_rack.cell(row=index, column=len(header) + 1, value="No Match")
            highlight_cell(ws_rack.cell(row=index, column=len(header) + 1), color="FFE78587")
            ws_rack_discrepancies.append(value) 

    ws_scanned['A1'] = "Scanned Glass"
    ws_scanned['A1'].font = Font(bold=True, size=14)
    for value in header:
        col_index = header.index(value) + 1
        ws_scanned.cell(row=2, column=col_index, value=value)  
        ws_scanned.cell(row=2, column=col_index).font = Font(bold=True)
        highlight_cell(ws_scanned.cell(row=2, column=col_index), color="FFADD8E6")
    ws_scanned.cell(row=2, column=len(header) + 1, value="Match Status (Compared to Racks)")
    ws_scanned.cell(row=2, column=len(header) + 1).font = Font(bold=True)
    highlight_cell(ws_scanned.cell(row=2, column=len(header) + 1), color="FFADD8E6")
    for index, value in enumerate(delivered.values(), start=3):
        if matching_piece_ids and value[0] in matching_piece_ids:
            for col_index, cell_value in enumerate(value, start=1):
                ws_scanned.cell(row=index, column=col_index, value=cell_value)
                highlight_cell(ws_scanned.cell(row=index, column=col_index), color="FF98FB98")
            ws_scanned.cell(row=index, column=len(header) + 1, value="Match")
            highlight_cell(ws_scanned.cell(row=index, column=len(header) + 1), color="FF98FB98")
        else:
            for col_index, cell_value in enumerate(value, start=1):
                ws_scanned.cell(row=index, column=col_index, value=cell_value)
                highlight_cell(ws_scanned.cell(row=index, column=col_index), color="FFE78587")
            ws_scanned.cell(row=index, column=len(header) + 1, value="No Match")
            highlight_cell(ws_scanned.cell(row=index, column=len(header) + 1), color="FFE78587")
            ws_scanned_discrepancies.append(value)

    output_file = os.path.join(rack_folder_path, 'Comparison Report.xlsx')
    wb.save(output_file)
    return output_file

def compare_scanned_discrepancy_manifest_reports(scanned_discrepancy_file_path, manifest_file_path):

    manifested = get_manifest_glass(manifest_file_path)

    manifested_keys = [
        (unit[0], unit[2], unit[3])
        for unit in manifested
    ]
    
    manifested_counter = Counter(manifested_keys)

    wb = load_workbook(scanned_discrepancy_file_path, data_only=False, read_only=False)
    ws = wb["Scanned Discrepancies"]

    for row in ws.iter_rows(min_row=3):
        key = (
            row[1].value,
            row[3].value,
            row[4].value
        )

        matching_tally = manifested_counter.get(key, 0)

        if matching_tally == 0:
            result = "Not Manifested"
            ws.cell(row=row[0].row, column=len(row)).value = result
            for cell in row:
                highlight_cell(cell, color="FFE78587")
        elif matching_tally == 1:
            result = "Manifested"
            ws.cell(row=row[0].row, column=len(row)).value = result
            for cell in row:
                highlight_cell(cell, color="FF98FB98")
        else:
            result = "Multiple Manifested"
            ws.cell(row=row[0].row, column=len(row)).value = result
            for cell in row:
                highlight_cell(cell, color="FFFFDFBF")

    wb.save(scanned_discrepancy_file_path)

def compare_rack_discrepancy_manifest_reports(rack_discrepancy_file_path, manifest_file_path):
    
    manifested = get_manifest_glass(manifest_file_path)

    manifested_keys = [
        (unit[0], unit[2], unit[3])
        for unit in manifested
    ]
    
    manifested_counter = Counter(manifested_keys)

    wb = load_workbook(rack_discrepancy_file_path, data_only=False, read_only=False)
    ws = wb["Rack Discrepancies"]

    for row in ws.iter_rows(min_row=3):
        key = (
            row[1].value,
            row[3].value,
            row[4].value
        )

        matching_tally = manifested_counter.get(key, 0)

        if matching_tally == 0:
            result = "Not Manifested"
            for cell in row:
                highlight_cell(cell, color="FFE78587")
        elif matching_tally == 1:
            result = "Manifested"
            for cell in row:
                highlight_cell(cell, color="FF98FB98")
        else:
            result = "Multiple Manifested"
            for cell in row:
                highlight_cell(cell, color="FFFFDFBF")

        ws.cell(row=row[0].row, column=len(row)).value = result

    wb.save(rack_discrepancy_file_path)

def run_rack_scanned_compare(rack_folder_path, scanned_file_path, runs):

    # if delivery_location == "OOT" and runs == ["All"]:
    #     runs = ["8310-Oamaru", "8311-Timaru", "8312-Ashburton", "8327-Chch To Nel Brn", "8329-Chch To Dun Brn"]
    # elif delivery_location == "Local" and runs == ["All"]:
    #     runs = ["8301: Christchurch 1", "8302: Christchurch 2", "8303: Christchurch 3"]
    # elif delivery_location == "All" and runs == ["All"]:
    #     runs = ["8310-Oamaru", "8311-Timaru", "8312-Ashburton", "8327-Chch To Nel Brn", "8329-Chch To Dun Brn", "8301: Christchurch 1", "8302: Christchurch 2", "8303: Christchurch 3"]
    # else:
    #     pass

    # parsed_folder_path = generate_rack_folder(rack_summaries_folder, customer_group_file, target_date, delivery_location, runs)
    # if parsed_folder_path == None:
    #     return None
    # else:
    rack_vs_scanned_comparison_report = compare_rack_scanned_reports(rack_folder_path, scanned_file_path, runs)

    return rack_vs_scanned_comparison_report



            