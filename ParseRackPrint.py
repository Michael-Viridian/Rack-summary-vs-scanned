# =============================================================================
# Module imports
# =============================================================================

import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook

# =============================================================================
# Function imports
# =============================================================================

from customer_patterns import _normalize_keywords, deliver_group_key

# =============================================================================
# Helper functions
# =============================================================================

def prompt_str(prompt, default=None):
    while True:
        raw = input(f"{prompt}" + (f" [{default}]" if default is not None else "") + ": ").strip()
        if raw == "" and default is not None:
            return default
        elif raw != "":
            return raw
        
# ----- robust date parser accepting multiple common formats -----
def _parse_date_any(date_str: str) -> datetime:
    """Try multiple common date formats and normalize to a datetime."""
    if date_str is None:
        raise ValueError("Date string is None")

    s = str(date_str).strip()

    # Common formats to try (ordered by likelihood / minimal ambiguity)
    candidates = [
        "%d/%m/%Y", "%d/%m/%y",
        "%d-%m-%Y", "%d-%m-%y",
        "%d.%m.%Y", "%d.%m.%y",
        "%m/%d/%Y", "%m/%d/%y",  # US
        "%m-%d-%Y", "%m-%d-%y",
        "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
        "%d %b %Y", "%d %B %Y",
        "%b %d, %Y", "%B %d, %Y",
        "%d %b %y", "%d %B %y",
        "%b %d, %y", "%B %d, %y",
    ]

    last_error = None
    for fmt in candidates:
        try:
            return datetime.strptime(s, fmt)
        except ValueError as e:
            last_error = e

    # Fallback heuristic: disambiguate dd/mm vs mm/dd
    for sep in ("/", "-", "."):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3 and all(p.isdigit() for p in parts):
                a, b, c = parts

                def _norm_year(y):
                    if len(y) == 2:
                        yy = int(y)
                        return 2000 + yy if yy <= 69 else 1900 + yy
                    return int(y)

                # Try D/M/Y
                try:
                    day = int(a)
                    month = int(b)
                    year = _norm_year(c)
                    return datetime(year, month, day)
                except Exception:
                    pass

                # Try M/D/Y
                try:
                    month = int(a)
                    day = int(b)
                    year = _norm_year(c)
                    return datetime(year, month, day)
                except Exception:
                    pass

    raise ValueError(f"Unrecognized date format for '{date_str}'. Last parsing error: {last_error}")

# =============================================================================
# Main function
# =============================================================================

def generate_rack_summary_file(folder_path):

    # Define the header for the output Excel file
    header = ['Piece ID', 'Order No', 'Deliver To', 'Product', 'Size', 'Rack No']

    # Define the custom order of columns to keep from the input text files
    custom_order = [0, 1, 2, 5, 6]

    output_file = os.path.join(folder_path, 'Rack_Summary.xlsx')
    wb = Workbook()
    ws = wb.active

    # Write the header row
    ws.append(header)

    for file in os.listdir(folder_path):
        if file.endswith('.txt'):
            rack_no = file.split('_')[0]
            file_path = os.path.join(folder_path, file)

            with open(file_path, 'r') as f:
                lines = f.readlines()

            for line in lines:
                if line[0].isdigit():
                    cleaned_line = line.strip()
                    parts = cleaned_line.split('\t')
                    kept_parts = [parts[i] for i in custom_order]
                    kept_parts.append(rack_no)  # Add the rack number to the end of the row
                    ws.append(kept_parts)

    wb.save(output_file)

    return output_file, header

def generate_rack_folder(source_folder_path, dest_folder_path=None, customer_group_file=None):

    customer_group_wb = load_workbook(customer_group_file, data_only=False) if customer_group_file else None
    customer_group_ws = customer_group_wb.active if customer_group_wb else None

    customer_names = [cell.value for cell in customer_group_ws['A'][1:]] if customer_group_ws else []
    patterns = _normalize_keywords(customer_names)

    target_datetime = _parse_date_any(prompt_str("Enter the target date (DDMMYY)"))
    start_datetime = target_datetime - timedelta(hours=9, minutes=30)

    local_or_OOT = prompt_str("Local or OOT?")

    if local_or_OOT == "OOT":
        end_datetime = target_datetime + timedelta(hours=4)
    else:
        end_datetime = target_datetime + timedelta(hours=14, minutes=30)

    for file in os.listdir(source_folder_path):
        if file.endswith('.txt'):
            file_path = Path(source_folder_path) / file
            file_stats = file_path.stat()
            file_modified_datetime = datetime.fromtimestamp(file_stats.st_mtime)
            if start_datetime <= file_modified_datetime <= end_datetime:

                with open(file_path, 'r') as f:
                    lines = f.readlines()

                for line in lines:
                    if line[0].isdigit():
                        cleaned_line = line.strip()
                        parts = cleaned_line.split('\t')
                        customer = deliver_group_key(parts[2], patterns)[0]

                        if customer in [pattern['label'] for pattern in patterns]:
                            if dest_folder_path is None:
                                dest_folder_path = os.path.join(source_folder_path, f"{local_or_OOT}_racks_{target_datetime.strftime('%d%m%y')}_del")
                            os.makedirs(dest_folder_path, exist_ok=True)
                            src_file_path = os.path.join(source_folder_path, file)
                            dest_file_path = os.path.join(dest_folder_path, file)
                            shutil.copy(src_file_path, dest_file_path)
                        break

