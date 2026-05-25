# =============================================================================
# Module imports
# =============================================================================

import os
from openpyxl import load_workbook, Workbook

# =============================================================================
# Function imports
# =============================================================================

from ParseRackPrint import generate_rack_summary_file, generate_rack_folder

# =============================================================================
# Constants
# =============================================================================

rack_prints_folder = r'C:\Users\MEVERETT\OneDrive - Viridian Glass Limited Partnership\Test - RackPrints'
delivered_glass_path = r'C:\Users\MEVERETT\OneDrive - Viridian Glass Limited Partnership\Test - RackPrints\Delivered_Glass.xlsx'

# =============================================================================
# Helper functions
# =============================================================================

def get_delivered_glass(delivered_glass_path):

    # Load workbook and first sheet
    wb = load_workbook(delivered_glass_path, data_only=False, read_only=False)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]

    col_index = {header: i for i, header in enumerate(headers)}

    new_order = ['Item Reference', 'Order No', 'Delivery Address', 'Product', 'Specification', 'Rack Number']

    new_data = []
    for row in ws.iter_rows(values_only=True):
        new_data.append([row[col_index[h]] for h in new_order])

    wb.remove(ws)
    ws = wb.create_sheet("Reordered")  

    # Write rearranged data
    for row in new_data:
        ws.append(row)

    delivered = {}

    for row in ws.iter_rows(min_row=3):

        piece_id = row[0].value  # Column A (index 0)

        row_data = []
        if piece_id is not None:
            for cell in row:
                row_data.append(cell.value)
            delivered[piece_id] = row_data

    return delivered

def get_rack_summary_glass():

    rack_summary_path, header = generate_rack_summary_file(rack_prints_folder)

    # Load workbook and first sheet
    wb = load_workbook(rack_summary_path, data_only=False, read_only=False)
    ws = wb.active

    rack_summaries = {}

    for row in ws.iter_rows(min_row=3):

        piece_id = row[0].value  # Column A (index 0)

        row_data = []
        if piece_id is not None:
            for cell in row:
                row_data.append(cell.value)
            rack_summaries[piece_id] = row_data

    return rack_summaries, header

# =============================================================================
# Main function
# =============================================================================

def compare_reports():

    rack_summary_glass, header = get_rack_summary_glass()
    delivered = get_delivered_glass(delivered_glass_path)

    matching_keys = set(rack_summary_glass.keys()) & set(delivered.keys())
    matching_values = [rack_summary_glass.get(key) for key in matching_keys]

    delivered_only_keys = set(delivered.keys()) - set(rack_summary_glass.keys())
    delivered_only_values = [delivered.get(key) for key in delivered_only_keys]

    rack_summary_only_keys = set(rack_summary_glass.keys()) - set(delivered.keys())
    rack_summary_only_values = [rack_summary_glass.get(key) for key in rack_summary_only_keys]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rack Summaries vs Delivered"

    ws['A1'] = "Matching Units"
    for value in header:
        col_index = header.index(value) + 1
        ws.cell(row=2, column=col_index, value=value)   
    for index, value in enumerate(matching_values, start=3):
        for col_index, cell_value in enumerate(value, start=1):
            ws.cell(row=index, column=col_index, value=cell_value)

    delivered_start_row = len(matching_values) + 4
    ws[f'A{delivered_start_row - 1}'] = "Delivered Only"

    # Header
    for col, value in enumerate(header, start=1):
        ws.cell(row=delivered_start_row, column=col, value=value)

    # Data (start AFTER header)
    for r, value in enumerate(delivered_only_values, start=delivered_start_row + 1):
        for c, cell_value in enumerate(value, start=1):
            ws.cell(row=r, column=c, value=cell_value)

    rack_summary_start_row = (
        delivered_start_row
        + len(delivered_only_values)
        + 2
    )
    ws[f'A{rack_summary_start_row - 1}'] = "Rack Summaries Only"

    # Header
    for col, value in enumerate(header, start=1):
        ws.cell(row=rack_summary_start_row, column=col, value=value)

    # Data
    for r, value in enumerate(rack_summary_only_values, start=rack_summary_start_row + 1):
        for c, cell_value in enumerate(value, start=1):
            ws.cell(row=r, column=c, value=cell_value)

    output_file = os.path.join(rack_prints_folder, 'Comparison Report.xlsx')
    wb.save(output_file)
    return output_file

# compare_reports()

generate_rack_folder(rack_prints_folder, customer_group_file=r"C:\Users\MEVERETT\OneDrive - Viridian Glass Limited Partnership\Projects\Truck Planning Automation\Unit-Optimisation\customer_groups.xlsx")


            